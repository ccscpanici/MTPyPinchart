import openpyxl
from utils import output
import time
import os
import shutil

test_file = os.getcwd() +  "\\Klondike Feta BF Pin Chart_StarTurn610.xlsm"

def get_sheet_data(sheet):
    rows = []
    row_index = 1
    for row in sheet.iter_rows():
        _cells = []
        column_index = 1

        for cell in row:

            # get check if cell is merged
            _cell_merged = False
            
            # get the row span for the cell
            _row_span = 1

            # get the column span for the cell
            _col_span = 1

            # instantiates the cell data dictionary
            _cell_dict = {}

            _cell_dict['row'] = cell.row
            _cell_dict['column'] = cell.column
            _cell_dict['value'] = cell.value
            _cell_dict['row_span'] = 1
            _cell_dict['column_span'] = 1
            _cell_dict['merged_cell'] = False
            _cell_dict['cell'] = cell

            try:
                _cell_dict['encoding'] = cell.encoding
            except:
                pass
            # end try

            for a_range in sheet.merged_cells.ranges:
                if cell == a_range.start_cell:
                    _cell_dict['merged_cell'] = True
                    _cell_dict['row_span'] = a_range.size['rows']
                    _cell_dict['column_span'] = a_range.size['columns']
                # end if
            # end for

            _cells.append(_cell_dict)
            column_index = column_index + 1
        # end for
        
        rows.append(_cells)
        row_index = row_index + 1
    # end for
    return rows
# end get_sheet_data

def get_xl_data(a_file):

    # get the file extension
    _extension = a_file.split(".")[-1]

    # get the destination directory
    _dest_dir = os.getcwd()

    # get the temp file name
    _temp_file = _dest_dir + "\\tmp_file." + _extension 
    
    # copy the file to a temporary location
    shutil.copy(a_file, _temp_file)

    # open up the temporary file
    wb = openpyxl.load_workbook(_temp_file, read_only=False, data_only=False, keep_links=False)
    sheet_data = {}
    sheets = wb.sheetnames
    for a_sheet in sheets:
        _sheet = wb[a_sheet]
        sheet_data[a_sheet] = get_sheet_data(_sheet)
    # end for

    # close the workbook (temporary)
    wb.close()

    # delete the temporary file
    os.remove(_temp_file)

    return sheet_data
# end get_xl_data
